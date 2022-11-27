from django.shortcuts import render,redirect
from app01 import models
from app01.utils.pagnation import Pagination
from openpyxl import load_workbook
# Create your views here.
# 部门管理
def depart_list(request):
    """展示部门列表"""
    # for i in range(50):
    #     models.Department.objects.create(title='网络部')
    data_dict = {}
    value = request.GET.get('q', '')
    if value:
        # name__contains要变
        data_dict['title__contains'] = value
    # models.Admin.objects.filter中Admin要变
    queryset = models.Department.objects.filter(**data_dict).order_by('id')
    # queryset=models.Department.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'value':value
    }
    return render(request,'depart_list.html',context)
def depart_add(request):
    """添加部门"""
    if request.method == 'GET':
        return render(request,'depart_add.html')
    title=request.POST.get('title')
    if title != '':
        models.Department.objects.create(title=title)
    return redirect('/depart/list/')
def depart_delete(request):
    """删除部门"""
    nid=request.GET.get('nid')
    models.Department.objects.filter(id=nid).delete()
    return redirect('/depart/list/')
def depart_edit(request):
    """编辑部门"""
    nid = request.GET.get('nid')
    if request.method == 'GET':
        obj = models.Department.objects.filter(id=nid).first()
        return render(request,'depart_edit.html',{'obj':obj.title})
    else:
        title = request.POST.get('title')
        # 暂时修改不了
        models.Department.objects.filter(id=nid).update(title=title)
        return redirect('/depart/list/')
def depart_upload(request):
    file_object = request.FILES.get('exc')
    if not file_object:
        return render(request,'depart_list.html',{'obj':'未选择文件'})
    # f = open('a1.png', mode='wb')
    # f = open(file_object.name, mode='wb')
    # for chunk in file_object.chunks():
    #     f.write(chunk)
    # f.close()

    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        exists = models.Department.objects.filter(title=text).exists()
        if not exists:
            models.Department.objects.create(title=text)
    return redirect('/depart/list/')